import glob
import dicom
import openpyxl
import json
import os
import sys


#Get dicom files
def get_dicom_files(paths):
    file_paths = glob.glob(paths)
    return file_paths


#Get "always wrong" patients
def get_always_wrong(json_file):
    data = json.load(open(json_file))
    wrongs = data['Statistics']['Always wrong']
    #Return dict
    return wrongs


#Get "always right" patients
def get_always_right(json_file):
    data = json.load(open(json_file))
    rights = data['Statistics']['Always right']
    #Return dict
    return rights


#Get "mostly wrong" patients
def get_mostly_wrong(json_file):
    data = json.load(open(json_file))
    wrongs = data['Statistics']['Mostly wrong']
    #Return dict
    return wrongs


#Get "mostly right" patients
def get_mostly_right(json_file):
    data = json.load(open(json_file))
    rights = data['Statistics']['Mostly right']
    #Return dict
    return rights


#Build Excel to analyse
def build_xls(xlsx_file, dicom_paths, sem_xlsx_file, wrongs=[None, None], rights=[None, None]):
    sem_file = openpyxl.load_workbook(sem_xlsx_file)
    sem_sheet = sem_file.active

    if os.path.exists(xlsx_file):
        os.remove(xlsx_file)
    wb = openpyxl.Workbook()
    ws = wb.active

    for i in range(len(dicom_paths)):
        data = dicom.read_file(dicom_paths[i])
        r = i + 1
        c = 0

        #Patient ID
        c += 1
        patient = data[0x10, 0x10][:]

        if r == 1:
            ws.cell(row=r, column=c, value='Patient_ID')
        else:
            ws.cell(row=r, column=c, value=patient)

        #Patient always or mostly wrong or right?
        if len(wrongs) == len(rights):
            if len(wrongs) == 2:
                c += 1
                if r == 1:
                    if wrongs[1] in ['Always', 'always']:
                        ws.cell(row=r, column=c, value='Always...')
                    elif wrongs[1] in ['Mostly', 'mostly']:
                        ws.cell(row=r, column=c, value='Mostly...')
                    elif wrongs[1] not in ['Always', 'always', 'Mostly', 'mostly']:
                        raise IOError('wrongs[1] and rights[1] must both be "Always" or "Mostly"')
                else:
                    if patient in wrongs[0]:
                        ws.cell(row=r, column=c, value='wrong')
                    elif patient in rights[0]:
                        ws.cell(row=r, column=c, value='right')
            else:
                raise IOError('wrongs and rights take exactly 2 arguments, {} given'.format(len(wrongs)))
        else:
            raise IOError('wrongs and rights must be same length, and both "Always" or both "Mostly"')

        #Patient sex
        c += 1
        if r == 1:
            ws.cell(row=r, column=c, value='Sex')
        else:
            ws.cell(row=r, column=c, value=data[0x10, 0x40].value)

        #Patient age
        c += 1
        if r == 1:
            ws.cell(row=r, column=c, value='Age')
        else:
            for i in range(1, sem_sheet.max_column):
                if sem_sheet.cell(row=1, column=i).value == 'Age':
                    column_index = i
                    break
            for i in range(1, sem_sheet.max_row):
                if sem_sheet.cell(row=i, column=1).value == patient:
                    ws.cell(row=r, column=c, value=sem_sheet.cell(row=i, column=column_index).value)
                    break

        #Location
        c += 1
        if r == 1:
            ws.cell(row=r, column=c, value='Tumour location')
        else:
            for i in range(1, sem_sheet.max_column):
                if sem_sheet.cell(row=1, column=i).value == 'Localisation':
                    column_index = i
                    break
            for i in range(1, sem_sheet.max_row):
                if sem_sheet.cell(row=i, column=1).value == patient:
                    ws.cell(row=r, column=c, value=sem_sheet.cell(row=i, column=column_index).value)
                    break

        #Depth
        c += 1
        if r == 1:
            ws.cell(row=r, column=c, value='Tumour depth')
        else:
            for i in range(1, sem_sheet.max_column):
                if sem_sheet.cell(row=1, column=i).value == 'Depth_Melissa':
                    column_index = i
                    break
            for i in range(1, sem_sheet.max_row):
                if sem_sheet.cell(row=i, column=1).value == patient:
                    ws.cell(row=r, column=c, value=sem_sheet.cell(row=i, column=column_index).value)
                    break

        #Manufacturer
        c += 1
        if r == 1:
            ws.cell(row=r, column=c, value='Manufacturer')
        else:
            ws.cell(row=r, column=c, value=data[0x08, 0x70].value)

        #Scanner model
        c += 1
        if r == 1:
            ws.cell(row=r, column=c, value='Model')
        else:
            ws.cell(row=r, column=c, value=data[0x08, 0x1090].value)

        #Institution name
        c += 1
        if r == 1:
            ws.cell(row=r, column=c, value='Institution')
        else:
            try:
                ws.cell(row=r, column=c, value=data[0x08, 0x80].value)
            except KeyError:
                ws.cell(row=r, column=c, value='-')

        #Station name
        c += 1
        if r == 1:
            ws.cell(row=r, column=c, value='Station')
        else:
            try:
                ws.cell(row=r, column=c, value=data[0x08, 0x1010].value)
            except KeyError:
                ws.cell(row=r, column=c, value='-')

        #Slice thickness
        c += 1
        if r == 1:
            ws.cell(row=r, column=c, value='Slice thickness')
        else:
            ws.cell(row=r, column=c, value=data[0x18, 0x50].value)

        #Repetition time
        c += 1
        if r == 1:
            ws.cell(row=r, column=c, value='RT')
        else:
            ws.cell(row=r, column=c, value=data[0x18, 0x80].value)

        #Echo time
        c += 1
        if r == 1:
            ws.cell(row=r, column=c, value='ET')
        else:
            ws.cell(row=r, column=c, value=data[0x18, 0x81].value)

        #Magnetic field strength
        c += 1
        if r == 1:
            ws.cell(row=r, column=c, value='B0')
        else:
            ws.cell(row=r, column=c, value=data[0x18, 0x87].value)

        #Spacing between slices
        c += 1
        if r == 1:
            ws.cell(row=r, column=c, value='Slice spacing')
        else:
            ws.cell(row=r, column=c, value=data[0x18, 0x88].value)

        #Smalles image pixel value
        c += 1
        if r == 1:
            ws.cell(row=r, column=c, value='Smallest pixel value')
        #else:
            #ws.cell(row=r, column=c, value=data[0x28, 0x0106].value)

        #Largest image pixel value
        c += 1
        if r == 1:
            ws.cell(row=r, column=c, value='Largest pixel value')
        #else:
            #ws.cell(row=r, column=c, value=data[0x28, 0x0107].value)

        wb.save(xlsx_file)
    print('file written to {}'.format(xlsx_file))
    return


def main(run_folder):
    #Input parameters
    sem_xlsx = 'Radiomics_melissa.xlsx'

    #Origins
    location = os.path.dirname(os.path.realpath(__file__))
    if 'wkessels' in location:
        dicom_paths = '/archive/wkessels/input/Lipo/*/*/*/*.dcm'
        json_file = '/archive/wkessels/output/{}/performance_all_0.json'.format(run_folder)
        xlsx_file = '/archive/wkessels/output/{}/analyse_out_semantics.xlsx'.format(run_folder)
        sem_xlsx_file = '/archive/wkessels/input/Patientdata/{}'.format(sem_xlsx)
    elif 'wouter' in location:
        dicom_paths = '/home/wouter/Documents/BEP_Wouter/Lipo/*/*/*/*.dcm'
        json_file = '/home/wouter/Documents/BEP_Wouter/Output_WORC/{}/performance_all_0.json'.format(run_folder)
        xlsx_file = '/home/wouter/Documents/BEP_Wouter/Output_WORC/{}/analyse_out_semantics.xlsx'.format(run_folder)
        sem_xlsx_file = '/home/wouter/Documents/BEP_Wouter/Patientdata/{}'.format(sem_xlsx)

    #Execute
    always_wrong = get_always_wrong(json_file)
    always_right = get_always_right(json_file)
    mostly_wrong = get_mostly_wrong(json_file)
    mostly_right = get_mostly_right(json_file)
    dicom_files = get_dicom_files(dicom_paths)
    build_xls(xlsx_file, dicom_files, sem_xlsx_file, wrongs=[mostly_wrong, 'mostly'], rights=[mostly_right, 'mostly'])


if __name__ == '__main__':
    if len(sys.argv) == 2:
        main(sys.argv[1])
    elif len(sys.argv) == 1:
        raise IOError('No input argument given.')
    else:
        raise IOError('main takes only 1 argument, {} given'.format(len(sys.argv)))
